"""
Write Charge Summary and Supervision Matrix to an Excel workbook.
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import Config

# Color palette
_HEADER_BG = "1F3864"   # dark navy
_HEADER_FG = "FFFFFF"
_SUBHEAD_BG = "D6E4F0"
_ALT_BG = "F2F7FB"
_TOTAL_BG = "E2EFDA"    # light green for total rows
_MD_BG = "EAF0FF"       # light blue for MD group headers


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_charge_summary(ws, summary: pd.DataFrame, config: Config, date_range: str, source_name: str):
    arial = "Arial"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Charge Summary by Supervising MD — {date_range}"
    ws["A1"].font = Font(name=arial, bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="left")

    meta_lines = [
        f"Date of Service: {date_range}  |  Source: {source_name}",
        "wRVU Source: CMS Physician Fee Schedule RVU26A, effective January 1, 2026",
    ]
    if config.excluded_signers:
        meta_lines.append(f"Excluded Signers (omitted from all output): {'; '.join(config.excluded_signers)}")
    overrides = []
    if config.reclassify_as_supervising_md:
        overrides.append(f"Reclassified as Supervising MDs: {', '.join(config.reclassify_as_supervising_md)}")
    if config.add_to_app_list:
        overrides.append(f"Added to APP list: {', '.join(config.add_to_app_list)}")
    if config.omit:
        overrides.append(f"Omitted: {', '.join(config.omit)}")
    if overrides:
        meta_lines.append("Manual Overrides — " + "  |  ".join(overrides))

    for i, line in enumerate(meta_lines, start=2):
        ws.merge_cells(f"A{i}:E{i}")
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = Font(name=arial, size=9, italic=True, color="555555")

    current_row = len(meta_lines) + 2
    ws.cell(row=current_row, column=1, value="Section A: wRVU Summary by Supervising MD")
    ws.cell(row=current_row, column=1).font = Font(name=arial, bold=True, size=11)
    current_row += 1

    for md, grp in summary.groupby("supervising_md", sort=True):
        md_total = grp["total_wrvu"].sum()

        # MD name row
        ws.merge_cells(f"A{current_row}:E{current_row}")
        cell = ws.cell(row=current_row, column=1, value=str(md))
        cell.font = Font(name=arial, bold=True, size=10)
        cell.fill = PatternFill("solid", start_color=_MD_BG)
        current_row += 1

        # Column headers
        headers = ["CPT Code", "Description", "Count", "wRVU/Unit", "Total wRVU"]
        for col_i, hdr in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_i, value=hdr)
            cell.font = Font(name=arial, bold=True, size=9, color=_HEADER_FG)
            cell.fill = PatternFill("solid", start_color=_HEADER_BG)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _thin_border()
        current_row += 1

        for row_i, (_, data_row) in enumerate(grp.iterrows()):
            bg = _ALT_BG if row_i % 2 == 0 else "FFFFFF"
            vals = [
                data_row["cpt"],
                data_row["description"],
                data_row["count"],
                data_row["wrvu_per_unit"],
                data_row["total_wrvu"],
            ]
            for col_i, val in enumerate(vals, 1):
                cell = ws.cell(row=current_row, column=col_i, value=val)
                cell.font = Font(name=arial, size=9)
                cell.fill = PatternFill("solid", start_color=bg)
                cell.border = _thin_border()
                if col_i in (3, 4, 5):
                    cell.alignment = Alignment(horizontal="right")
                    if col_i in (4, 5):
                        cell.number_format = "0.00"
            current_row += 1

        # TOTAL row
        for col_i in range(1, 6):
            cell = ws.cell(row=current_row, column=col_i)
            cell.fill = PatternFill("solid", start_color=_TOTAL_BG)
            cell.font = Font(name=arial, bold=True, size=9)
            cell.border = _thin_border()
        ws.cell(row=current_row, column=1, value="TOTAL")
        ws.cell(row=current_row, column=3, value=int(grp["count"].sum()))
        ws.cell(row=current_row, column=5, value=round(md_total, 2))
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=current_row, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=current_row, column=5).number_format = "0.00"
        current_row += 2  # blank row between MDs

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12


def _write_supervision_matrix(ws, matrix: pd.DataFrame, date_range: str, config: Config):
    arial = "Arial"
    apps = [c for c in matrix.columns if c != "Total wRVU (Sup MD)"]
    all_cols = apps + ["Total wRVU (Sup MD)"]
    mds = [r for r in matrix.index if r != "APP Total wRVU Signed"]

    excluded_str = "; ".join(config.excluded_signers) if config.excluded_signers else "None"

    ws.merge_cells(f"A1:{get_column_letter(len(all_cols) + 1)}1")
    ws["A1"] = f"Supervision Matrix — Inpatient APPs by Supervising MD (% of each APP's total wRVUs signed)"
    ws["A1"].font = Font(name=arial, bold=True, size=12)

    ws.merge_cells(f"A2:{get_column_letter(len(all_cols) + 1)}2")
    ws["A2"] = (
        f"% = APP's wRVUs attributed to this Supervising MD ÷ APP's total wRVUs across all true Sup MDs. "
        f"Excluded signers: {excluded_str}. Date of Service: {date_range}."
    )
    ws["A2"].font = Font(name=arial, size=9, italic=True, color="555555")

    header_row = 3
    ws.cell(row=header_row, column=1, value="Supervising MD")
    ws.cell(row=header_row, column=1).font = Font(name=arial, bold=True, size=9, color=_HEADER_FG)
    ws.cell(row=header_row, column=1).fill = PatternFill("solid", start_color=_HEADER_BG)
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=header_row, column=1).border = _thin_border()

    for col_i, app in enumerate(all_cols, 2):
        cell = ws.cell(row=header_row, column=col_i, value=app)
        cell.font = Font(name=arial, bold=True, size=9, color=_HEADER_FG)
        cell.fill = PatternFill("solid", start_color=_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _thin_border()
        ws.row_dimensions[header_row].height = 40

    for row_i, md in enumerate(mds):
        excel_row = header_row + 1 + row_i
        bg = _ALT_BG if row_i % 2 == 0 else "FFFFFF"

        cell = ws.cell(row=excel_row, column=1, value=md)
        cell.font = Font(name=arial, size=9)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.border = _thin_border()

        for col_i, col_name in enumerate(all_cols, 2):
            val = matrix.loc[md, col_name]
            cell = ws.cell(row=excel_row, column=col_i)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.border = _thin_border()
            cell.font = Font(name=arial, size=9)
            cell.alignment = Alignment(horizontal="right")
            if pd.notna(val) and val != 0:
                cell.value = float(val)
                if col_name == "Total wRVU (Sup MD)":
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "0.0%"

    # APP totals row
    totals_row = header_row + 1 + len(mds)
    ws.cell(row=totals_row, column=1, value="APP Total wRVU Signed")
    ws.cell(row=totals_row, column=1).font = Font(name=arial, bold=True, size=9)
    ws.cell(row=totals_row, column=1).fill = PatternFill("solid", start_color=_TOTAL_BG)
    ws.cell(row=totals_row, column=1).border = _thin_border()

    for col_i, col_name in enumerate(all_cols, 2):
        val = matrix.loc["APP Total wRVU Signed", col_name]
        cell = ws.cell(row=totals_row, column=col_i)
        cell.fill = PatternFill("solid", start_color=_TOTAL_BG)
        cell.font = Font(name=arial, bold=True, size=9)
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="right")
        if pd.notna(val) and val is not None:
            cell.value = float(val)
            cell.number_format = "#,##0.00"

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col_i in range(2, len(all_cols) + 2):
        ws.column_dimensions[get_column_letter(col_i)].width = 10
    # Last col wider for Total wRVU
    ws.column_dimensions[get_column_letter(len(all_cols) + 1)].width = 14

    ws.freeze_panes = "B4"


def write_workbook(
    output_path: str,
    summary: pd.DataFrame,
    matrix: pd.DataFrame,
    config: Config,
    date_range: str,
    source_name: str,
):
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Charge Summary"

    ws_matrix = wb.create_sheet("Supervision Matrix")

    _write_charge_summary(ws_summary, summary, config, date_range, source_name)
    _write_supervision_matrix(ws_matrix, matrix, date_range, config)

    wb.save(output_path)
    return output_path
